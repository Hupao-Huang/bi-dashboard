import openpyxl
import pymysql
import sys
import os

DB_CONFIG = {
    'host': '127.0.0.1',
    'port': 3306,
    'user': 'root',
    'password': 'Hch123456',
    'database': 'bi_dashboard',
    'charset': 'utf8mb4',
}

SHEET_DEPT_MAP = {
    '考核利润汇总表': '汇总',
    '1、电商': '电商', '电商': '电商',
    '2、社媒': '社媒', '社媒': '社媒',
    '3、线下': '线下', '线下': '线下',
    '4、分销': '分销', '分销': '分销',
    '5、私域': '私域', '私域': '私域',
    '6、国际零售业务': '国际零售',
    '7、即时零售': '即时零售',
    '8、糙有力量': '糙有力量',
    '中台部门': '中台',
}

CATEGORY_KEYWORDS = [
    ('GMV', 'GMV'),
    ('营业收入', '收入'),
    ('营业成本', '成本'),
    ('仓储物流费用', '成本'),
    ('营业毛利', '毛利'),
    ('销售费用', '销售费用'),
    ('运营利润', '运营利润'),
    ('管理费用', '管理费用'),
    ('研发费用', '研发费用'),
    ('利润总额', '利润总额'),
    ('营业利润', '利润总额'),
    ('营业外收入', '营业外'),
    ('营业外支出', '营业外'),
    ('净利润', '净利润'),
]

FILES = {
    r'C:\Users\Administrator\Desktop\财务报表\25年12月财务报表.xlsx': 2025,
    r'C:\Users\Administrator\Desktop\财务报表\26年1-33月财务报表.xlsx': 2026,
}

def import_file(conn, fpath, year):
    wb = openpyxl.load_workbook(fpath, data_only=True)
    total = 0

    for sheet_name in wb.sheetnames:
        dept = SHEET_DEPT_MAP.get(sheet_name)
        if not dept:
            continue

        ws = wb[sheet_name]
        max_col = ws.max_column or 30

        # Find month columns (row 2 is header)
        month_cols = {}
        for ci in range(1, min(max_col + 1, 30)):
            val = ws.cell(2, ci).value
            if val is None:
                continue
            val = str(val).strip()
            for m in range(1, 13):
                if val == f'{m}月':
                    month_cols[ci] = m
                    break

        if not month_cols:
            continue

        current_category = ''
        sort_order = 0
        sheet_count = 0

        for ri in range(3, min((ws.max_row or 80) + 1, 80)):
            subject = ws.cell(ri, 1).value
            if subject is None:
                continue
            subject = str(subject).strip()
            if not subject or subject == '项目':
                continue

            for kw, cat in CATEGORY_KEYWORDS:
                if kw in subject:
                    current_category = cat
                    break

            sort_order += 1

            for ci, month in month_cols.items():
                val = ws.cell(ri, ci).value
                if val is None:
                    continue
                if isinstance(val, str):
                    val = val.strip()
                    if val in ('', '#DIV/0!', '#REF!', '-'):
                        continue
                    try:
                        val = float(val)
                    except ValueError:
                        continue

                if not isinstance(val, (int, float)) or val == 0:
                    continue

                with conn.cursor() as cur:
                    cur.execute('''
                        INSERT INTO finance_report (year, month, department, subject, subject_category, sort_order, amount)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE
                            subject_category=VALUES(subject_category), sort_order=VALUES(sort_order), amount=VALUES(amount)
                    ''', (year, month, dept, subject, current_category, sort_order, round(val, 2)))
                sheet_count += 1
                total += 1

        conn.commit()
        print(f'  [{sheet_name}] {dept} → {sheet_count} 条')

    return total

def main():
    conn = pymysql.connect(**DB_CONFIG)

    # Clear existing data
    with conn.cursor() as cur:
        cur.execute('DELETE FROM finance_report')
    conn.commit()
    print('已清空旧数据')

    for fpath, year in FILES.items():
        if not os.path.exists(fpath):
            print(f'文件不存在: {fpath}')
            continue
        print(f'处理 {os.path.basename(fpath)} (年份: {year})')
        count = import_file(conn, fpath, year)
        print(f'  共导入 {count} 条')

    # Summary
    with conn.cursor() as cur:
        cur.execute('SELECT year, COUNT(*), COUNT(DISTINCT department), COUNT(DISTINCT month) FROM finance_report GROUP BY year')
        for row in cur.fetchall():
            print(f'  {row[0]}年: {row[1]}条, {row[2]}个部门, {row[3]}个月')

    conn.close()

if __name__ == '__main__':
    main()
